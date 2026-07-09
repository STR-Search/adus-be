"""Backfill legacy underwritings from the 'Underwritten Properties' Google Sheet export.

Reads the XLSX export of the underwriting Google Sheet and loads each deal
into iron_bank verbatim (no recalculation), marked with source='legacy_sheet'
and its sheet tab/link number so the team can verify rows against the sheet.

Deal sources inside the workbook:
- Main_Sheet rows (header row 4) keyed by the 'Link' column
- Delete_Properties rows (header row 1) for deals already removed
- one tab per deal, named by its number ('233'..'1937'), parsed by section
  labels rather than fixed coordinates because the template shifted rows
  across versions.

Idempotent: deals whose sheet_number already exists in the DB are skipped
(and a partial unique index enforces this DB-side). Use --update to
delete-and-reinsert a range after parser fixes or a fresh export.

Usage:
  python scripts/backfill_legacy_underwritings.py --dry-run
  python scripts/backfill_legacy_underwritings.py --range 1558:1937
  python scripts/backfill_legacy_underwritings.py --range 1558:1600 --update
"""

import argparse
import asyncio
import json
import re
import sys
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

sys.path.append(str(Path(__file__).resolve().parent.parent))

DEFAULT_XLSX = Path(__file__).resolve().parent.parent / "Underwritten Properties.xlsx"
REPORT_PATH = Path(__file__).resolve().parent / "backfill_report.json"

MAIN_SHEET = "Main_Sheet"
DELETE_SHEET = "Delete_Properties"
CLIENT_SHOWN_SHEET = "ClientShown_Properties"

# Status values as plain strings (mirroring app.iron_bank.enums.DealStatus and
# UnderwritingSource) so --dry-run never imports app.* — the app package pulls
# in config/DB at import time, which requires DATABASE_URL.
LEGACY_SOURCE = "legacy_sheet"
# Blank sheet statuses (no summary row, or an empty status cell) become this
# fixed DealStatus value.
NO_STATUS = "Previously Underwritten - No Status"

# Sheet status label -> deal_status enum value.
STATUS_MAP: dict[str, str] = {
    "template generated": "template_generated",
    "analyst started": "analyst_started",
    "analyst completed": "analyst_completed",
    "delete - zillow": "delete_zillow",
    "delete zillow": "delete_zillow",
    "delete - deal": "delete_deal",
    "delete deal": "delete_deal",
    "delete": "delete_deal",
    "maybe": "maybe",
    "maybe (save for later)": "maybe",
    "re-forecast revenue": "re_forecast_revenue",
    "re forecast revenue": "re_forecast_revenue",
    "awaiting realtor details": "awaiting_realtor_details",
    "waiting on realtor": "awaiting_realtor_details",
    "present to clients": "present_to_clients",
    "client under contract": "client_under_contract",
    "training deal": "training_deal",
    "training deal for onboarding": "training_deal",
    "floorplan/ video needed": "awaiting_realtor_details",
    "taylor review needed": "analyst_completed",
}

# Sheet analyst labels that don't match "First LastInitial" against users.users.
# Values are matched against "first_name last_name" (lowercased). Extend as the
# unmatched-name warnings surface real cases.
NICKNAME_OVERRIDES: dict[str, str] = {
    # "rizz (ahmed)": "ahmed <last name>",
}


# ---------------------------------------------------------------------------
# value normalization
# ---------------------------------------------------------------------------


def to_decimal(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip().replace("$", "").replace(",", "").replace("%", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def to_int(value: Any) -> int | None:
    dec = to_decimal(value)
    return int(dec) if dec is not None else None


def to_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        dt = value
    else:
        text = str(value).strip()
        if not text:
            return None
        try:
            dt = datetime.fromisoformat(text)
        except ValueError:
            return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def to_bool(value: Any) -> bool:
    return str(value).strip().lower() == "true"


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    # the sheet uses the literal string 'None' as its empty marker
    if not text or text.lower() == "none":
        return None
    return text


def jsonable(value: Any) -> Any:
    """Recursively converts Decimals to floats for JSONB columns."""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, dict):
        return {k: jsonable(v) for k, v in value.items()}
    if isinstance(value, list):
        return [jsonable(v) for v in value]
    return value


# ---------------------------------------------------------------------------
# summary rows (Main_Sheet / Delete_Properties)
# ---------------------------------------------------------------------------

# Both tabs share column names for the fields we need; only header row differs.
SUMMARY_FIELDS = {
    "Deal_Status": "raw_status",
    "Property Address": "property_address",
    "City": "city",
    "ST": "state",
    "Analyst": "analyst_name",
    "Approved By": "approver_name",
    "Approved Time": "deal_approved",
    "PP": "purchase_price",
    "Cash Needed": "total_oop",
    "PRR": "prr",
    "Low": "low_gross_revenue",
    "Mid": "mid_gross_revenue",
    "High": "high_gross_revenue",
    "L": "l_cash_on_cash",
    "M": "m_cash_on_cash",
    "H": "h_cash_on_cash",
    "Date_Added": "deal_added",
    "Bedrooms": "sleep_capacity",
    "Turnkey?": "turnkey",
    "Property Pending": "property_pending",
    "Loom_Vid": "loom_vid",
    "Notes": "notes",
    "Link": "link",
}


def index_summary_rows(ws, header_row: int) -> dict[int, dict[str, Any]]:
    """Returns {sheet_number: raw summary dict} for one summary tab."""
    rows = ws.iter_rows(min_row=header_row, values_only=True)
    headers = next(rows, None)
    if headers is None:
        return {}
    col_for = {
        name: idx for idx, name in enumerate(headers) if name in SUMMARY_FIELDS
    }
    address_col = col_for.get("Property Address")
    result: dict[int, dict[str, Any]] = {}
    for row_number, row in enumerate(rows, start=header_row + 1):
        # read-only mode trims trailing empty cells, so rows vary in length
        raw = {
            SUMMARY_FIELDS[name]: row[idx] if idx < len(row) else None
            for name, idx in col_for.items()
        }
        link = to_int(raw.get("link"))
        if link is not None and _summary_has_content(raw):
            # where this row's address cell sits, e.g. "I5" — the tracking tabs
            # hyperlink the listing URL on the address text
            if address_col is not None:
                raw["_address_ref"] = f"{_column_letter(address_col)}{row_number}"
            result[link] = raw
    return result


def _column_letter(index: int) -> str:
    """0-based column index -> spreadsheet letters (0 -> A, 26 -> AA)."""
    letters = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def _summary_has_content(raw: dict[str, Any]) -> bool:
    """Main_Sheet pre-numbers empty rows for future deals (Link filled,
    everything else blank) — those placeholders are not deals."""
    return any(
        clean_text(raw.get(field)) is not None
        for field in (
            "property_address",
            "raw_status",
            "purchase_price",
            "total_oop",
            "deal_added",
            "deal_approved",
            "loom_vid",
            "notes",
        )
    )


# ---------------------------------------------------------------------------
# deal tab parsing (label-anchored)
# ---------------------------------------------------------------------------

# column indexes in the grid (0-based): B=1, C=2, D=3, E=4, F=5, H=7, I=8,
# J=9, K=10, L=11, M=12


def _cell(grid: list[tuple], row: int, col: int) -> Any:
    if 0 <= row < len(grid) and 0 <= col < len(grid[row]):
        return grid[row][col]
    return None


def _find_row(grid: list[tuple], col: int, prefix: str) -> int | None:
    prefix = prefix.lower()
    for idx in range(len(grid)):
        value = _cell(grid, idx, col)
        if value is not None and str(value).strip().lower().startswith(prefix):
            return idx
    return None


def _parse_purchase_details(grid, warnings: list[str]) -> dict[str, Any]:
    start = _find_row(grid, 1, "purchase details")
    if start is None:
        warnings.append("section not found: Purchase Details")
        return {}
    labels = {
        "purchase price": ("purchase_price", 3),
        "down payment": ("down_payment", 3),
        "loan amount": ("loan_amount", 3),
        "interest rate": ("interest_rate", 3),
        "mortgage years": ("mortgage_years", 3),
        "closing costs": ("closing_costs", 3),
    }
    pct_labels = {"down payment": "down_payment_pct", "closing costs": "closing_costs_pct"}
    details: dict[str, Any] = {}
    for idx in range(start + 1, min(start + 10, len(grid))):
        label = clean_text(_cell(grid, idx, 1))
        if label is None:
            continue
        key = label.lower()
        for prefix, (field, col) in labels.items():
            if key.startswith(prefix):
                value = to_decimal(_cell(grid, idx, col))
                if value is not None:
                    details[field] = value
                if prefix in pct_labels:
                    pct = to_decimal(_cell(grid, idx, 2))
                    if pct is not None:
                        details[pct_labels[prefix]] = pct
                break
    return details


def _parse_construction_loan(grid) -> dict[str, Any] | None:
    """v4 tabs only: 'Loan Amount / Int Rate (Annual) / Term (Years)' in K/L/M."""
    for idx in range(len(grid)):
        if (
            clean_text(_cell(grid, idx, 10)) == "Loan Amount"
            and clean_text(_cell(grid, idx, 11)) is not None
            and "int rate" in str(_cell(grid, idx, 11)).lower()
        ):
            amount = to_decimal(_cell(grid, idx + 1, 10))
            if amount and amount > 0:
                return {
                    "loan_amount": amount,
                    "interest_rate": to_decimal(_cell(grid, idx + 1, 11)),
                    "term_years": to_decimal(_cell(grid, idx + 1, 12)),
                }
            return None
    return None


def _parse_optimization_items(grid, warnings: list[str]) -> list[dict[str, Any]]:
    # 'optim' catches both 'Optimization List (Estimate)' and the oldest
    # template's typo 'Optimzation List (Rough Estimate)'
    start = _find_row(grid, 4, "optim")
    if start is None:
        warnings.append("section not found: Optimization List")
        return []
    items: list[dict[str, Any]] = []
    for idx in range(start + 1, len(grid)):
        label = clean_text(_cell(grid, idx, 4))
        if label is None:
            continue
        if label.lower().startswith("total"):
            break
        price = to_decimal(_cell(grid, idx, 5))
        items.append({"category": label, "total_price": price})
    return items


def _parse_operating_expenses(grid, warnings: list[str]) -> list[dict[str, Any]]:
    start = _find_row(grid, 7, "operating expenses (opex)")
    if start is None:
        warnings.append("section not found: Operating Expenses (OPEX)")
        return []
    expenses: list[dict[str, Any]] = []
    for idx in range(start + 1, len(grid)):
        label = clean_text(_cell(grid, idx, 7))
        if label is None:
            continue
        lowered = label.lower()
        if lowered.startswith("total operating expenses"):
            break
        if lowered.startswith("disclaimer"):
            continue
        amount = to_decimal(_cell(grid, idx, 8))
        if amount is None and lowered in ("other", "misc"):
            continue
        expenses.append({"expense_name": label, "monthly_amount": amount})
    return expenses


def _parse_cleaning_cost(grid) -> dict[str, Any] | None:
    for idx in range(len(grid)):
        if (
            clean_text(_cell(grid, idx, 10)) == "Cleaning Cost"
            and clean_text(_cell(grid, idx, 11)) == "# of Turns"
        ):
            return {
                "cleaning_cost": to_decimal(_cell(grid, idx + 1, 10)),
                "turns_per_month": to_decimal(_cell(grid, idx + 1, 11)),
            }
    return None


def _parse_taxes(grid) -> dict[str, Any] | None:
    start = _find_row(grid, 1, "taxes")
    if start is None:
        return None
    labels = {
        "land assumptions": "land_assumptions_pct",
        "improvement basis": "improvement_basis",
        "estimated short life assets": "estimated_short_life_assets",
        "bonus amount": "bonus_amount_pct",
        "tax rate": "tax_rate_pct",
        "y1 loss from depreciation": "y1_loss_from_depreciation",
        "tax savings": "tax_savings",
    }
    taxes: dict[str, Any] = {}
    for idx in range(start + 1, min(start + 10, len(grid))):
        label = clean_text(_cell(grid, idx, 1))
        if label is None:
            continue
        key = label.lower()
        for prefix, field in labels.items():
            if key.startswith(prefix):
                value = to_decimal(_cell(grid, idx, 2))
                if value is not None:
                    taxes[field] = value
                break
    return taxes or None


# Sheet row label (column B) -> per-scenario key, matching the shape
# SaveUnderwritingService persists for app-created rows so the FE reads
# legacy and app underwritings identically. Values live in D/E/F = low/mid/high.
_SCENARIO_ROWS = {
    "forecasted revenue": "forecasted_revenue",
    "operating expenses (annual": "operating_expenses_annual",
    "co-hosting fee": "co_hosting_fee",
    "net operating income": "net_operating_income",
    "debt service (annual)": "debt_service_annual",
    "annual free cash flow": "annual_free_cash_flow",
}


def _parse_forecasted_revenue(grid) -> dict[str, Any] | None:
    anchor = _find_row(grid, 1, "forecasted revenue")
    if anchor is None:
        return None
    scenarios: dict[str, dict[str, Any]] = {"low": {}, "mid": {}, "high": {}}
    co_hosting_fee_pct = None
    for idx in range(anchor, min(anchor + 12, len(grid))):
        label = clean_text(_cell(grid, idx, 1))
        if label is None:
            continue
        key = label.lower()
        for prefix, field in _SCENARIO_ROWS.items():
            if key.startswith(prefix):
                for col, scenario in ((3, "low"), (4, "mid"), (5, "high")):
                    value = to_decimal(_cell(grid, idx, col))
                    if value is not None:
                        scenarios[scenario][field] = value
                if field == "co_hosting_fee":
                    co_hosting_fee_pct = to_decimal(_cell(grid, idx, 2))
                break
    if not any(scenarios.values()):
        return None
    result: dict[str, Any] = {"scenarios": scenarios}
    if co_hosting_fee_pct is not None:
        result["co_hosting_fee_pct"] = co_hosting_fee_pct
    return result


def _parse_y1_coc(grid) -> dict[str, Any] | None:
    """'Year 1 Cash on Cash, including Tax Savings' block (label in column D,
    Low/Mid/High header row, then the percentage values)."""
    for idx in range(len(grid)):
        label = clean_text(_cell(grid, idx, 3))
        if label is None or not label.lower().startswith("year 1 cash on cash"):
            continue
        for j in range(idx + 1, min(idx + 4, len(grid))):
            low = to_decimal(_cell(grid, j, 3))
            if low is not None:
                return {
                    "low_pct": low,
                    "mid_pct": to_decimal(_cell(grid, j, 4)),
                    "high_pct": to_decimal(_cell(grid, j, 5)),
                }
    return None


def _parse_comp_set(grid) -> list[dict[str, Any]]:
    header = None
    for idx in range(len(grid)):
        if clean_text(_cell(grid, idx, 7)) == "Listing URL":
            header = idx
            break
    if header is None:
        return []
    comps: list[dict[str, Any]] = []
    for idx in range(header + 1, len(grid)):
        url = clean_text(_cell(grid, idx, 7))
        if url is None:
            break
        comps.append(
            {
                "listing_url": url,
                "revenue": to_decimal(_cell(grid, idx, 8)),
                "bedrooms": to_int(_cell(grid, idx, 9)),
                "sleeps": to_int(_cell(grid, idx, 10)),
            }
        )
    return comps


def _parse_analyst_notes(grid) -> str | None:
    idx = _find_row(grid, 10, "analyst notes")
    if idx is None:
        return None
    return clean_text(_cell(grid, idx + 1, 10))


def _parse_prepared_by(grid) -> str | None:
    idx = _find_row(grid, 4, "prepared by")
    if idx is None:
        return None
    return clean_text(_cell(grid, idx, 5))


def _parse_property_url(grid) -> str | None:
    idx = _find_row(grid, 4, "property url")
    if idx is None:
        return None
    value = clean_text(_cell(grid, idx, 5))
    if value is not None and value.lower().startswith("http"):
        return value
    return None


def parse_deal_tab(grid: list[tuple]) -> dict[str, Any]:
    """Parses one deal tab's cell grid into child-record inputs + warnings."""
    warnings: list[str] = []
    purchase_details = _parse_purchase_details(grid, warnings)
    construction_loan = _parse_construction_loan(grid)
    if construction_loan:
        purchase_details["construction_loan"] = construction_loan
    return {
        "purchase_details": purchase_details,
        "cleaning_cost": _parse_cleaning_cost(grid),
        "forecasted_revenue": _parse_forecasted_revenue(grid),
        "y1_coc_incl_tax_savings": _parse_y1_coc(grid),
        "taxes": _parse_taxes(grid),
        "optimization_items": _parse_optimization_items(grid, warnings),
        "operating_expenses": _parse_operating_expenses(grid, warnings),
        "comp_set": _parse_comp_set(grid),
        "analyst_notes": _parse_analyst_notes(grid),
        "prepared_by": _parse_prepared_by(grid),
        "listing_url": _parse_property_url(grid),
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# deal assembly
# ---------------------------------------------------------------------------


def map_deal_status(raw_status: Any) -> str:
    label = clean_text(raw_status)
    if label is None:
        return NO_STATUS
    status = STATUS_MAP.get(label.lower())
    if status is None:
        raise ValueError(f"unmapped sheet deal status: {label!r}")
    return status


def build_deal(
    sheet_number: int,
    summary: dict[str, Any] | None,
    tab: dict[str, Any] | None,
    listing_url: str | None = None,
) -> dict[str, Any]:
    """Assembles the repository.create() inputs for one deal."""
    warnings: list[str] = []
    notes: list[str] = []
    underwriting: dict[str, Any] = {
        "source": LEGACY_SOURCE,
        "sheet_number": sheet_number,
        "is_automated": False,
    }
    if listing_url:
        underwriting["listing_url"] = listing_url

    if summary is None:
        warnings.append("no summary row in any tracking tab (deal tab only)")
        underwriting["deal_status"] = NO_STATUS
    else:
        underwriting["deal_status"] = map_deal_status(summary.get("raw_status"))
        address = clean_text(summary.get("property_address"))
        underwriting.update(
            property_address=address,
            street=address.split(",")[0].strip() if address else None,
            city=clean_text(summary.get("city")),
            state=clean_text(summary.get("state")),
            purchase_price=to_decimal(summary.get("purchase_price")),
            total_oop=to_decimal(summary.get("total_oop")),
            prr=to_decimal(summary.get("prr")),
            low_gross_revenue=to_decimal(summary.get("low_gross_revenue")),
            mid_gross_revenue=to_decimal(summary.get("mid_gross_revenue")),
            high_gross_revenue=to_decimal(summary.get("high_gross_revenue")),
            l_cash_on_cash=to_decimal(summary.get("l_cash_on_cash")),
            m_cash_on_cash=to_decimal(summary.get("m_cash_on_cash")),
            h_cash_on_cash=to_decimal(summary.get("h_cash_on_cash")),
            deal_added=to_datetime(summary.get("deal_added")),
            deal_approved=to_datetime(summary.get("deal_approved")),
            sleep_capacity=to_int(summary.get("sleep_capacity")),
            turnkey=to_bool(summary.get("turnkey")),
            property_pending=to_bool(summary.get("property_pending")),
            loom_vid=clean_text(summary.get("loom_vid")),
        )
        pp = underwriting.get("purchase_price")
        oop = underwriting.get("total_oop")
        if pp and oop and pp > 0:
            underwriting["budget_to_pp"] = oop / pp
        # note carries only the sheet's own Notes column
        if clean_text(summary.get("notes")):
            notes.append(str(summary["notes"]).strip())

    detail: dict[str, Any] = {}
    taxes = None
    optimization_items: list[dict[str, Any]] = []
    operating_expenses: list[dict[str, Any]] = []
    comp_set: list[dict[str, Any]] = []

    if tab is None:
        warnings.append("no deal tab in workbook (summary row only)")
    else:
        warnings.extend(tab["warnings"])
        if tab["purchase_details"]:
            # JSONB columns can't take Decimals; numeric fidelity is preserved
            # in the typed columns, so floats are fine inside the JSON payloads
            detail["purchase_details"] = jsonable(tab["purchase_details"])
            if underwriting.get("purchase_price") is None:
                underwriting["purchase_price"] = tab["purchase_details"].get(
                    "purchase_price"
                )
        if tab["cleaning_cost"]:
            detail["cleaning_cost"] = jsonable(tab["cleaning_cost"])
        if tab["forecasted_revenue"]:
            detail["forecasted_revenue"] = jsonable(tab["forecasted_revenue"])
        if tab["y1_coc_incl_tax_savings"]:
            detail["y1_coc_incl_tax_savings"] = jsonable(
                tab["y1_coc_incl_tax_savings"]
            )
        if tab["analyst_notes"]:
            detail["analyst_notes"] = tab["analyst_notes"]
        if tab["listing_url"] and not underwriting.get("listing_url"):
            underwriting["listing_url"] = tab["listing_url"]
        taxes = tab["taxes"]
        optimization_items = tab["optimization_items"]
        operating_expenses = tab["operating_expenses"]
        comp_set = tab["comp_set"]

    analyst_name = clean_text(summary.get("analyst_name")) if summary else None
    if analyst_name is None and tab is not None:
        analyst_name = tab["prepared_by"]
    approver_name = clean_text(summary.get("approver_name")) if summary else None

    return {
        "sheet_number": sheet_number,
        "underwriting": underwriting,
        "detail": detail or None,
        "taxes": taxes,
        "optimization_items": optimization_items,
        "operating_expenses": operating_expenses,
        "comp_set": comp_set,
        "analyst_name": analyst_name,
        "approver_name": approver_name,
        "notes": notes,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# workbook reading
# ---------------------------------------------------------------------------


def _sheet_targets(z) -> dict[str, str]:
    """Tab name -> worksheet xml path inside the xlsx."""
    workbook = z.read("xl/workbook.xml").decode()
    wb_rels = z.read("xl/_rels/workbook.xml.rels").decode()
    rel_target = {
        m.group(1): m.group(2)
        for m in re.finditer(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', wb_rels)
    }
    return {
        m.group(1): rel_target[m.group(2)]
        for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="(rId\d+)"', workbook)
        if m.group(2) in rel_target
    }


def _external_hyperlinks(z, target: str) -> dict[str, str]:
    """Cell ref -> external URL for one worksheet file."""
    sheet_xml = z.read(f"xl/{target}").decode()
    links = re.findall(r'<hyperlink r:id="(rId\d+)" ref="([A-Z]+\d+)"', sheet_xml)
    if not links:
        return {}
    try:
        sheet_rels = z.read(
            f"xl/worksheets/_rels/{target.split('/')[-1]}.rels"
        ).decode()
    except KeyError:
        return {}
    targets = {
        m.group(1): m.group(2)
        for m in re.finditer(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', sheet_rels)
    }
    return {
        ref: targets[rid]
        for rid, ref in links
        if rid in targets and targets[rid].startswith("http")
    }


def extract_listing_urls(
    path: Path, tracking_rows: dict[str, dict[int, dict[str, Any]]]
) -> dict[int, str]:
    """Listing links live as hyperlinks in two places, neither visible to
    openpyxl's read-only mode, so they're pulled from the xlsx XML:
    - the PROPERTY URL / address cell (E/F, top rows) of each deal tab
    - the Property Address cell of each tracking-tab row (e.g. Main_Sheet I5)
    Precedence: deal tab > Main_Sheet > ClientShown > Delete_Properties."""
    import zipfile

    urls: dict[int, str] = {}
    with zipfile.ZipFile(path) as z:
        sheet_target = _sheet_targets(z)

        # lowest priority first; later writes win
        for tab in (DELETE_SHEET, CLIENT_SHOWN_SHEET, MAIN_SHEET):
            target = sheet_target.get(tab)
            if target is None or tab not in tracking_rows:
                continue
            by_ref = _external_hyperlinks(z, target)
            for link, raw in tracking_rows[tab].items():
                url = by_ref.get(raw.get("_address_ref", ""))
                if url:
                    urls[link] = url

        for name, target in sheet_target.items():
            if not re.fullmatch(r"\d+", name.strip()):
                continue
            by_ref = _external_hyperlinks(z, target)
            for ref, url in sorted(by_ref.items()):
                if re.fullmatch(r"[EF][1-6]", ref):
                    urls[int(name)] = url
                    break
    return urls


def read_workbook(path: Path) -> dict[str, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    main_rows = index_summary_rows(wb[MAIN_SHEET], header_row=4)
    deleted_rows = (
        index_summary_rows(wb[DELETE_SHEET], header_row=1)
        if DELETE_SHEET in wb.sheetnames
        else {}
    )
    client_shown_rows = (
        index_summary_rows(wb[CLIENT_SHOWN_SHEET], header_row=1)
        if CLIENT_SHOWN_SHEET in wb.sheetnames
        else {}
    )
    tabs: dict[int, list[tuple]] = {}
    for name in wb.sheetnames:
        if re.fullmatch(r"\d+", name.strip()):
            tabs[int(name)] = [
                row for row in wb[name].iter_rows(max_row=80, values_only=True)
            ]
    wb.close()
    # When a link appears in several tracking tabs:
    # Main_Sheet > ClientShown_Properties > Delete_Properties.
    summaries = {**deleted_rows, **client_shown_rows, **main_rows}
    return {
        "summaries": summaries,
        "tabs": tabs,
        "listing_urls": extract_listing_urls(
            path,
            {
                MAIN_SHEET: main_rows,
                CLIENT_SHOWN_SHEET: client_shown_rows,
                DELETE_SHEET: deleted_rows,
            },
        ),
    }


# ---------------------------------------------------------------------------
# user matching
# ---------------------------------------------------------------------------


def build_user_matcher(users: list[Any]):
    """Matches sheet labels like 'Taylor J' / 'John B' to users.users ids."""
    by_key: dict[str, int] = {}
    for user in users:
        first = (user.first_name or "").strip().lower()
        last = (user.last_name or "").strip().lower()
        if not first:
            continue
        if last:
            by_key.setdefault(f"{first} {last}", user.id)
            by_key.setdefault(f"{first} {last[0]}", user.id)
        else:
            by_key.setdefault(first, user.id)

    def match(name: str | None) -> int | None:
        if not name:
            return None
        key = name.strip().lower()
        key = NICKNAME_OVERRIDES.get(key, key)
        return by_key.get(key)

    return match


def split_person_name(label: str) -> tuple[str, str | None]:
    """'Taylor J' -> ('Taylor', 'J'); 'Kevin' -> ('Kevin', None)."""
    first, _, last = label.strip().partition(" ")
    return first, (last.strip() or None)


def legacy_clerk_id(label: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", label.strip().lower()).strip("_")
    return f"legacy_{slug}"


# ---------------------------------------------------------------------------
# load
# ---------------------------------------------------------------------------


async def load_deals(deals: list[dict[str, Any]], update: bool) -> dict[str, Any]:
    from sqlalchemy import delete, select

    from app.core.database import AsyncSessionLocal
    from app.iron_bank.models import Underwriting
    from app.iron_bank.repositories.underwriting_repository import (
        UnderwritingRepository,
    )
    from app.users.models.user import User

    inserted, skipped, failed = [], [], []
    async with AsyncSessionLocal() as session:
        existing = set(
            (
                await session.execute(
                    select(Underwriting.sheet_number).where(
                        Underwriting.sheet_number.isnot(None)
                    )
                )
            )
            .scalars()
            .all()
        )
        users = (
            (await session.execute(select(User).where(User.is_deleted.isnot(True))))
            .scalars()
            .all()
        )
        match_user = build_user_matcher(users)
        created_users: dict[str, int] = {}
        repository = UnderwritingRepository(session)

        async def resolve_user(name: str | None) -> int | None:
            """Match an existing user, or create a placeholder to link to."""
            if not name:
                return None
            user_id = match_user(name)
            if user_id is not None:
                return user_id
            key = name.strip().lower()
            if key not in created_users:
                first, last = split_person_name(name)
                user = User(
                    clerk_id=legacy_clerk_id(name), first_name=first, last_name=last
                )
                session.add(user)
                # commit immediately so a later failed deal insert can't roll
                # the user back while its id stays cached in created_users
                await session.commit()
                created_users[key] = user.id
            return created_users[key]

        for deal in deals:
            number = deal["sheet_number"]
            if number in existing:
                if not update:
                    skipped.append(number)
                    continue
                await session.execute(
                    delete(Underwriting).where(
                        Underwriting.sheet_number == number,
                        Underwriting.source == LEGACY_SOURCE,
                    )
                )
                await session.commit()

            underwriting = dict(deal["underwriting"])
            underwriting["analyst_id"] = await resolve_user(deal["analyst_name"])
            underwriting["approver_id"] = await resolve_user(deal["approver_name"])
            if deal["notes"]:
                underwriting["note"] = "\n".join(deal["notes"])

            try:
                created = await repository.create(
                    underwriting_data=underwriting,
                    detail_data=deal["detail"],
                    tax_data=deal["taxes"],
                    optimization_items=deal["optimization_items"],
                    operating_expenses=deal["operating_expenses"],
                    comp_set=deal["comp_set"],
                )
                inserted.append({"sheet_number": number, "id": created.id})
            except Exception as exc:  # keep going; report at the end
                failed.append({"sheet_number": number, "error": str(exc)})

    return {
        "inserted": inserted,
        "skipped": skipped,
        "failed": failed,
        "created_users": created_users,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def parse_range(text: str | None) -> tuple[int, int] | None:
    if text is None:
        return None
    low, _, high = text.partition(":")
    return int(low), int(high)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Backfill legacy underwritings from the Google Sheet XLSX export."
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help="Path to the 'Underwritten Properties' export.",
    )
    parser.add_argument(
        "--gsheet",
        type=str,
        default=None,
        help="Google Sheet spreadsheet id (cron mode; not implemented yet).",
    )
    parser.add_argument(
        "--range",
        type=str,
        default=None,
        help="Only process sheet numbers LOW:HIGH inclusive, e.g. 1558:1937.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and report only; no database access.",
    )
    parser.add_argument(
        "--update",
        action="store_true",
        help="Delete-and-reinsert deals that already exist (parser fixes, fresh export).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.gsheet is not None:
        sys.exit(
            "--gsheet mode is planned for the cron phase: it needs gspread and a "
            "GOOGLE_SERVICE_ACCOUNT_FILE the sheet is shared with. Use --xlsx for now."
        )
    if not args.xlsx.exists():
        sys.exit(f"XLSX not found: {args.xlsx}")

    bounds = parse_range(args.range)
    data = read_workbook(args.xlsx)
    numbers = sorted(set(data["summaries"]) | set(data["tabs"]))
    if bounds:
        numbers = [n for n in numbers if bounds[0] <= n <= bounds[1]]

    deals = [
        build_deal(
            n,
            data["summaries"].get(n),
            _parsed_tab(data, n),
            listing_url=data["listing_urls"].get(n),
        )
        for n in numbers
    ]

    if args.dry_run:
        result = {"inserted": [], "skipped": [], "failed": [], "created_users": {}}
    else:
        result = asyncio.run(load_deals(deals, update=args.update))

    report = {
        "xlsx": str(args.xlsx),
        "range": args.range,
        "dry_run": args.dry_run,
        "deals_found": len(deals),
        "inserted": len(result["inserted"]),
        "skipped_existing": len(result["skipped"]),
        "created_placeholder_users": result["created_users"],
        "failed": result["failed"],
        "deals_with_warnings": [
            {"sheet_number": deal["sheet_number"], "warnings": deal["warnings"]}
            for deal in deals
            if deal["warnings"]
        ],
    }
    REPORT_PATH.write_text(json.dumps(report, indent=2, default=str))

    print(
        json.dumps(
            {k: v for k, v in report.items() if k != "deals_with_warnings"},
            indent=2,
            default=str,
        )
    )
    print(
        f"{len(report['deals_with_warnings'])} deals with warnings "
        f"-> {REPORT_PATH}"
    )


def _parsed_tab(data: dict[str, Any], number: int) -> dict[str, Any] | None:
    grid = data["tabs"].get(number)
    return parse_deal_tab(grid) if grid is not None else None


if __name__ == "__main__":
    main()
